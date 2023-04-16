import openpyxl
import xlwings
import pandas as pd
import datetime as dt
import numpy as np
import matplotlib.pyplot as plt
from numpy import var
from matplotlib.ticker import ScalarFormatter

excelfile = 'C:/Users/ArvindSanala/PycharmProjects/pythonProject/venv/20230213-Series Sion Complete Vehicle - Working Structure_A3.xlsx'
workbook = openpyxl.load_workbook(excelfile, data_only=False)
sheet = workbook.active
sheet.cell(row=1, column=28, value="Prognosis Weight")
mr = sheet.max_row  # Last row to add formula to
for row in sheet.iter_rows(min_col=28, max_col=28, min_row=2, max_row=mr):
    for cell in row:
        cr = cell.row  # Get the current row number to use in formula
        cell.value = f'=IF(ISNUMBER(N{cr})*(N{cr} <> 0), N{cr}, IF(ISNUMBER(Q{cr})*(Q{cr} <> 0), Q{cr}, IF(ISBLANK(N{cr})*ISBLANK(P{cr})*ISBLANK(Q{cr}), 0,P{cr})))'
workbook.save(excelfile)
excel_app = xlwings.App(visible=False)
excel_book = excel_app.books.open(excelfile)
excel_book.save()
excel_book.close()
excel_app.quit()
df2 = pd.DataFrame(pd.read_excel(excelfile))
rows = ["Instance Collaborative Space"]
columns = ["Enterprise Item Number"]
values = ["Weighed Weight (GRAM)", "Calculated Weight (GRAM)", "Estimated Weight (GRAM)", "Prognosis Weight"]
relavent = df2.loc[:, rows + values]
pf1 = df2.pivot_table(index='Instance Collaborative Space', aggfunc='sum',
                      values=['Weighed Weight (GRAM)', 'Calculated Weight (GRAM)', 'Estimated Weight (GRAM)',
                              'Prognosis Weight'], margins=True, margins_name='Total Car Weight', )
fig = plt.figure("Weight Estimation 02/06/2023")
fig.suptitle("Weight Progression")
ax1 = fig.add_axes([0.1, 0.5, 0.7, 0.4])
ax1.set_ylabel("Weight in Grams")
ax1.ticklabel_format(useOffset=False,style='plain')
# ax1.axvline(y=10000, color="black", linewidth=2)
final1 = df2.groupby('Instance Collaborative Space')[
    'Weighed Weight (GRAM)', 'Calculated Weight (GRAM)', 'Estimated Weight (GRAM)', 'Prognosis Weight'].sum().plot(
    kind="bar", ax=ax1)
pf1.to_excel("WeightInputReport20230213-KW 08.xlsx")
